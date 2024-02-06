import numpy as np
import pandas as pd #Leer excel
pd.options.display.max_columns = None #TODO: Buscar utilidad
import os
import pdfplumber
from openpyxl import load_workbook #Editar Excel
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from PyPDF2 import PdfReader #Leer PDFs
import re
import tabula


def extract_text_from_pdf(pdf_path):
    extracted_text = ""

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            extracted_text += page.extract_text()

    return extracted_text

class Cliente:
    def __init__(self, CUIT_PJ, Contribuyente, Responsable):
        self.CUIT_PJ = CUIT_PJ
        #self.período = período
        self.Contribuyente = Contribuyente
        #self.tipo = tipo
        #self.impuesto = impuesto
        self.Responsable = Responsable
        
class NominaLaboral:
    def __init__(self, juris_cod, juris_desc, imp_determinado, saldo_favor):
        self.juris_cod = juris_cod
        self.juris_desc = juris_desc
        self.imp_determinado = imp_determinado
        self.saldo_favor = saldo_favor

def estandarizarnombre(nombreCliente):
    nombreCliente = nombreCliente.replace("&", "")
    nombreCliente = nombreCliente.replace(".", "")
    nombreCliente = nombreCliente.replace("SRL", "")
    nombreCliente = nombreCliente.replace("SA", "")
    nombreCliente = nombreCliente.replace(" ", "")
    return nombreCliente.upper()    

def base_writer(hoja, index, col, value):
    celda = hoja[col + str(index)]
    celda.value = value


#Variable del Nombre del archivo
fileName = "NominaLaboral.xlsx"
#Variable del Nombre de la hoja
sheetName = 'BD'
#Variable de fecha de Declaraciones juradas a ejecutar
fecha = '11/2023'    

#Variable del Nombre del excel
#TODO Add excel path

#Apertura de excel con pandas para lectura
df = pd.read_excel(fileName, sheet_name=sheetName)

#Apertura de excel con openpyxl para escritura
book = load_workbook(fileName)

#Obtiene el nombre de todas las hojas del documento
nombres_hojas = book.sheetnames

#Asigna formato a las columnas con pandas
#df['Período'] = df['Período'].astype(str)
#df['Impuesto Determinado'] = df['Impuesto Determinado'].astype(str)
#df['Saldo a Favor'] = df['Saldo a Favor'].astype(str)  

#1. Leer el excel (hoja BD) y guardar lista de clase Cliente
#Validación de existencia de hoja BASE
if 'BD' in nombres_hojas:
    #Validación de existencia de hoja LOGS
    if 'LOGS' in nombres_hojas:
        hoja_logs = book['LOGS']
        #Borrado de datos de Logs
        hoja_logs.delete_cols(1)
    else:
        hoja_logs = book.create_sheet(title="LOGS")
        
     #Validación de existencia de hoja BASE
    if 'BASE_NOMINAS' in nombres_hojas:
        hoja_base = book['BASE_NOMINAS']
        #Borrado de datos de Logs
        hoja_base.delete_rows(2, hoja_base.max_row)
    
    else:
        hoja_base = book.create_sheet(title="BASE_NOMINAS")
    
    if 'BASE_DDJJ' in nombres_hojas:
        hoja_base2 = book['BASE_DDJJ']
        #Borrado de datos de Logs
        hoja_base2.delete_rows(3, hoja_base2.max_row)
    
    else:
        hoja_base2 = book.create_sheet(title="BASE")
        
    #Variable con la ruta de las carpetas de los clientes
    path_client_folders = '../'
    #Listar los nombre de las carpetas de los clientes
    client_folders = os.listdir(path_client_folders)
    
    #Variable de lista vacia de clientes
    clientes = []
    
    #Recorre las lineas del excel con pandas. i = numero de iteración; f = linea de cliente en el excel
    for i, f in df.iterrows():
        c = Cliente(f.CUIT_PJ, f.Contribuyente, f.Responsable)
        clientes.append(c)
        
    index = 0
    celda_index = 2
    celda_indexddjj = 2

    #Comienza a recorrer la lista de clientes a actualizar
    for c in clientes:
        index += 1
        año = fecha.split('/')[1]
        mes = fecha.split('/')[0]
        mes_año = f"{mes}"
        
        try: 
            pdf_path = ''
            for f in client_folders:
                if estandarizarnombre(f) == estandarizarnombre(c.Contribuyente):
                    pdf_path = str(f'..\{f}/{año}/{mes_año}')
                    break
                
           
            pdfs = os.listdir(pdf_path)
            
            exists_pdf = False
            for p in pdfs:
                #Se obtiene la celda de la Hoja Logs lista para escribir
                celda_logs = hoja_logs[f'A{index}']
                
            
                if ('F931 NOMINA' in p.upper() or 'NOMINA' in p.upper()) and 'ACUSE' not in p.upper():
                    exists_pdf = True
                
                    archivo_pdf = f'{pdf_path}/{p}'
                    tablas_pdf = tabula.read_pdf(archivo_pdf, pages='all', multiple_tables=True)
     
                    # Convertir las tablas a DataFrames de pandas
                    dataframes = [pd.DataFrame(tabla) for tabla in tablas_pdf]
                    
                    #Abrir edicion de excel
                    #TODO open function
                    
                    # Imprimir las tablas
                    for i, df in enumerate(dataframes):
                        if i % 2 == 1:
                            for col in range(1, 5):
                                name_column = 'nan'
                                index_name_column = -1
                                
                                while True:
                                    # Verificar la condición para determinar si continuar o salir del bucle
                                    if name_column != 'Apellido y Nombre' and name_column != 'nan':
                                        break  # Salir del bucle si la condición no se cumple
                                    
                                    index_name_column += 1
                                    # Código que se ejecutará al menos una vez
                                    name_column = ('nan' if pd.isna(df.iloc[index_name_column, 0]) else (df.iloc[index_name_column, 0] ))
                                    
                                modalidad = df.iloc[6 + index_name_column, col]
                                if modalidad == '99':
                                    cuil = df.columns[col]
                                    nombre = ''
                                    for row in range(0, index_name_column):
                                        nombre += ('' if pd.isna(df.iloc[row, col]) else (df.iloc[row, col] + ' '))
                                
                                    remuneracion_total = df.iloc[13 + index_name_column, col]

                                    #Se obtiene la celda de la Hoja Base lista para escribir
                                    base_writer(hoja_base, celda_index, 'A', fecha)
                                    base_writer(hoja_base, celda_index, 'B', c.Responsable)
                                    base_writer(hoja_base, celda_index, 'C', c.CUIT_PJ)
                                    base_writer(hoja_base, celda_index, 'D', c.Contribuyente)
                                    base_writer(hoja_base, celda_index, 'E', nombre)
                                    base_writer(hoja_base, celda_index, 'F', cuil)
                                    base_writer(hoja_base, celda_index, 'G', '99')
                                    base_writer(hoja_base, celda_index, 'H', remuneracion_total)
                
                                    celda_index = celda_index +1
                                    
                        
                                   #TODO write function
                if ('F931 DDJJ' in p.upper() or 'DDJJ' in p.upper()) and 'ACUSE' not in p.upper():
                    exists_pdf = True  
                    
                    text = extract_text_from_pdf(f'{pdf_path}/{p}')  
                     
                    Suma_de_Rem_1 = text.split('\nSuma de Rem. 1: ')[1].split('\nfalseado información que deba')[0]
                    Suma_de_Rem_2 = text.split('\nSuma de Rem. 2: ')[1].split('\ncontener esta declaración, siendo fiel\nDeclaración Jurada')[0]
                    Suma_de_Rem_3 = text.split('Suma de Rem. 3: ')[1].split('\nPesos con centavos expresión de la verdad.\n')[0]
                    Suma_de_Rem_4 = text.split('Suma de Rem. 4: ')[1].split('\nS.U.S.S.')[0]
                    Suma_de_Rem_5 = text.split('Suma de Rem. 5: ')[1].split('\nSuma de Rem. 6:')[0]
                    Suma_de_Rem_6 = text.split('Suma de Rem. 6: ')[1].split('\nApellido y Nombre o Razón Social: ')[0]
                    Suma_de_Rem_7 = text.split('Suma de Rem. 7: ')[1].split('\nVerificador:\n')[0]
                    Suma_de_Rem_8 = text.split('Suma de Rem. 8: ')[1].split('\nSuma de Rem. 9:')[0]
                    Suma_de_Rem_9 = text.split('Suma de Rem. 9: ')[1].split('\nSuma de Rem. 10:')[0]
                    Suma_de_Rem_10 = text.split('Suma de Rem. 10: ')[1].split('\nDomicilio Fiscal:')[0]
                    Contribuciones_de_Seguridad_Social = text.split ('Contribuciones de Seguridad Social')[1].split('-')[0]
                    Aportes_de_Obra_Social = text.split ('Aportes de Obra Social')[1].split('\n301')[0]
                    Aportes_de_Seguridad_Social = text.split ('Aportes de Seguridad Social')[1].split('- Vales Alimentarios')[0]
                    Contribuciones_de_Obra_Social = text.split ('Contribuciones de Obra Social')[1].split('- Seguro Colectivo')[0]
                    L_R_T = text.split ('L.R.T.')[1].split('\n352')[0]
                    Seguro_Colectivo_de_Vida_Obligatorio = text.split ('Seguro Colectivo de Vida Obligatorio')[1].split('\n935')[0]
                    #texto = Suma_de_Rem_1         
                    #patron = r"\d+\s(.+?)\s\d+\.\d+%"
                    #resultado = re.search(patron, texto)
                    #if resultado:
                    #    parte_deseada = resultado.group(1)


                    base_writer(hoja_base2, celda_indexddjj, 'A', fecha)
                    base_writer(hoja_base2, celda_indexddjj, 'B', c.Responsable)
                    base_writer(hoja_base2, celda_indexddjj, 'C', c.CUIT_PJ)
                    base_writer(hoja_base2, celda_indexddjj, 'D', c.Contribuyente)
                    base_writer(hoja_base2, celda_indexddjj, 'E', "Suma de Rem. 1")
                    base_writer(hoja_base2, celda_indexddjj, 'F',  Suma_de_Rem_1)
                    base_writer(hoja_base2, celda_indexddjj, 'G', "351")
                    base_writer(hoja_base2, celda_indexddjj, 'H', "Contribuciones de Seguridad Social")
                    base_writer(hoja_base2, celda_indexddjj, 'I', Contribuciones_de_Seguridad_Social)
                    celda_indexddjj += 1
                    base_writer(hoja_base2, celda_indexddjj, 'A', fecha)
                    base_writer(hoja_base2, celda_indexddjj, 'B', c.Responsable)
                    base_writer(hoja_base2, celda_indexddjj, 'C', c.CUIT_PJ)
                    base_writer(hoja_base2, celda_indexddjj, 'D', c.Contribuyente)
                    base_writer(hoja_base2, celda_indexddjj, 'E', "Suma de Rem. 2")
                    base_writer(hoja_base2, celda_indexddjj, 'F',  Suma_de_Rem_2)
                    base_writer(hoja_base2, celda_indexddjj, 'G', "302")
                    base_writer(hoja_base2, celda_indexddjj, 'H', "Aportes de Obra Social")
                    base_writer(hoja_base2, celda_indexddjj, 'I', Aportes_de_Obra_Social)
                    
                    celda_indexddjj += 1
                    base_writer(hoja_base2, celda_indexddjj, 'A', fecha)
                    base_writer(hoja_base2, celda_indexddjj, 'B', c.Responsable)
                    base_writer(hoja_base2, celda_indexddjj, 'C', c.CUIT_PJ)
                    base_writer(hoja_base2, celda_indexddjj, 'D', c.Contribuyente)
                    base_writer(hoja_base2, celda_indexddjj, 'E', "Suma de Rem. 3")
                    base_writer(hoja_base2, celda_indexddjj, 'F',  Suma_de_Rem_3)
                    base_writer(hoja_base2, celda_indexddjj, 'G', "301")
                    base_writer(hoja_base2, celda_indexddjj, 'H', "Aportes de Seguridad Social")
                    base_writer(hoja_base2, celda_indexddjj, 'I', Aportes_de_Seguridad_Social)
                    
                    celda_indexddjj += 1
                    base_writer(hoja_base2, celda_indexddjj, 'A', fecha)
                    base_writer(hoja_base2, celda_indexddjj, 'B', c.Responsable)
                    base_writer(hoja_base2, celda_indexddjj, 'C', c.CUIT_PJ)
                    base_writer(hoja_base2, celda_indexddjj, 'D', c.Contribuyente)
                    base_writer(hoja_base2, celda_indexddjj, 'E', "Suma de Rem. 4")
                    base_writer(hoja_base2, celda_indexddjj, 'F',  Suma_de_Rem_4)
                    base_writer(hoja_base2, celda_indexddjj, 'G', "352")
                    base_writer(hoja_base2, celda_indexddjj, 'H', "Contribuciones de Obra Social")
                    base_writer(hoja_base2, celda_indexddjj, 'I', Contribuciones_de_Obra_Social)
                    
                    celda_indexddjj += 1
                    base_writer(hoja_base2, celda_indexddjj, 'A', fecha)
                    base_writer(hoja_base2, celda_indexddjj, 'B', c.Responsable)
                    base_writer(hoja_base2, celda_indexddjj, 'C', c.CUIT_PJ)
                    base_writer(hoja_base2, celda_indexddjj, 'D', c.Contribuyente)
                    base_writer(hoja_base2, celda_indexddjj, 'E', "Suma de Rem. 5")
                    base_writer(hoja_base2, celda_indexddjj, 'F',  Suma_de_Rem_5)
                    base_writer(hoja_base2, celda_indexddjj, 'G', "312")
                    base_writer(hoja_base2, celda_indexddjj, 'H', "L.R.T.")
                    base_writer(hoja_base2, celda_indexddjj, 'I',  L_R_T)
                    
                    celda_indexddjj += 1
                    base_writer(hoja_base2, celda_indexddjj, 'A', fecha)
                    base_writer(hoja_base2, celda_indexddjj, 'B', c.Responsable)
                    base_writer(hoja_base2, celda_indexddjj, 'C', c.CUIT_PJ)
                    base_writer(hoja_base2, celda_indexddjj, 'D', c.Contribuyente)
                    base_writer(hoja_base2, celda_indexddjj, 'E', "Suma de Rem. 6")
                    base_writer(hoja_base2, celda_indexddjj, 'F',  Suma_de_Rem_6)
                    base_writer(hoja_base2, celda_indexddjj, 'G', "028")
                    base_writer(hoja_base2, celda_indexddjj, 'H', "Seguro Colectivo de Vida Obligatorio")
                    base_writer(hoja_base2, celda_indexddjj, 'I', Seguro_Colectivo_de_Vida_Obligatorio)
                    
                    celda_indexddjj += 1
                    base_writer(hoja_base2, celda_indexddjj, 'A', fecha)
                    base_writer(hoja_base2, celda_indexddjj, 'B', c.Responsable)
                    base_writer(hoja_base2, celda_indexddjj, 'C', c.CUIT_PJ)
                    base_writer(hoja_base2, celda_indexddjj, 'D', c.Contribuyente)
                    base_writer(hoja_base2, celda_indexddjj, 'E', "Suma de Rem. 7")
                    base_writer(hoja_base2, celda_indexddjj, 'F',  Suma_de_Rem_7)
                    
                    celda_indexddjj += 1
                    base_writer(hoja_base2, celda_indexddjj, 'A', fecha)
                    base_writer(hoja_base2, celda_indexddjj, 'B', c.Responsable)
                    base_writer(hoja_base2, celda_indexddjj, 'C', c.CUIT_PJ)
                    base_writer(hoja_base2, celda_indexddjj, 'D', c.Contribuyente)
                    base_writer(hoja_base2, celda_indexddjj, 'E', "Suma de Rem. 8")
                    base_writer(hoja_base2, celda_indexddjj, 'F',  Suma_de_Rem_8)
                
                    celda_indexddjj += 1
                    base_writer(hoja_base2, celda_indexddjj, 'A', fecha)
                    base_writer(hoja_base2, celda_indexddjj, 'B', c.Responsable)
                    base_writer(hoja_base2, celda_indexddjj, 'C', c.CUIT_PJ)
                    base_writer(hoja_base2, celda_indexddjj, 'D', c.Contribuyente)
                    base_writer(hoja_base2, celda_indexddjj, 'E', "Suma de Rem. 9")
                    base_writer(hoja_base2, celda_indexddjj, 'F',  Suma_de_Rem_9)
                    
                    celda_indexddjj += 1
                    base_writer(hoja_base2, celda_indexddjj, 'A', fecha)
                    base_writer(hoja_base2, celda_indexddjj, 'B', c.Responsable)
                    base_writer(hoja_base2, celda_indexddjj, 'C', c.CUIT_PJ)
                    base_writer(hoja_base2, celda_indexddjj, 'D', c.Contribuyente)
                    base_writer(hoja_base2, celda_indexddjj, 'E', "Suma de Rem. 10")
                    base_writer(hoja_base2, celda_indexddjj, 'F',  Suma_de_Rem_10)
                    
                    celda_indexddjj += 1
                
                    
            #Cerrar el archivo excel para terminar la edicion (close)
            #TODO close function
                                    
                                    

                

        except FileNotFoundError:
            celda_logs = hoja_logs[f'A{index}']
            celda_logs.value = f'El cliente {c.Contribuyentel} no se pudo encontrar dentro de la carpeta WNS'
        except IndexError:
            celda_logs = hoja_logs[f'A{index}']
            celda_logs.value = f'El pdf del cliente {c.Contribuyente} es incorrecto'
        except OSError:
            celda_logs = hoja_logs[f'A{index}']
            celda_logs.value = f'Error, la ruta es incorrecta para {c.Contribuyente}'
            
    book.save(fileName)
    book.close()

else:
    print("La hoja 'BD' no existe en el archivo.")